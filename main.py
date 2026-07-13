import asyncio
import csv
import html
import logging
import os
import re
import sqlite3
import tempfile
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from telegram import (
    BotCommand,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Update,
)
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# ============================================================
# KONFIGURASI
# ============================================================

BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
DB_PATH = os.getenv("DB_PATH", "data/surat_keluar.db").strip()
ALLOWED_USER_IDS_RAW = os.getenv("ALLOWED_USER_IDS", "").strip()
ADMIN_IDS_RAW = os.getenv("ADMIN_IDS", "").strip()
SEQUENCE_MODE = os.getenv("SEQUENCE_MODE", "global").strip().lower()
START_NUMBER = int(os.getenv("START_NUMBER", "1"))
APP_TIMEZONE = os.getenv("APP_TIMEZONE", "Asia/Jakarta").strip()

if SEQUENCE_MODE not in {"global", "per_classification"}:
    raise ValueError("SEQUENCE_MODE harus 'global' atau 'per_classification'.")

try:
    LOCAL_TZ = ZoneInfo(APP_TIMEZONE)
except Exception as exc:
    raise ValueError(f"APP_TIMEZONE tidak valid: {APP_TIMEZONE}") from exc


def parse_id_list(raw: str) -> set[int]:
    result: set[int] = set()
    if not raw:
        return result
    for item in raw.split(","):
        item = item.strip()
        if item:
            result.add(int(item))
    return result


ALLOWED_USER_IDS = parse_id_list(ALLOWED_USER_IDS_RAW)
ADMIN_IDS = parse_id_list(ADMIN_IDS_RAW)

# ============================================================
# KLASIFIKASI SURAT
# ============================================================

CLASSIFICATIONS = {
    "500.13": "PARIWISATA DAN EKONOMI KREATIF",
    "500.13.1": "Kebijakan di bidang Pariwisata dan Ekonomi Kreatif yang dilakukan oleh Pemerintah Daerah",

    "500.13.2": "Pengembangan Destinasi Wisata",
    "500.13.2.1": "Perancangan Destinasi dan Investasi Pariwisata",
    "500.13.2.2": "Pengembangan Daya Tarik Wisata",
    "500.13.2.3": "Industri Pariwisata",
    "500.13.2.4": "Pemberdayaan Masyarakat Destinasi Pariwisata",
    "500.13.2.5": "Pengembangan Wisata Minat Khusus, Konvensi, Insentif, dan Event",

    "500.13.3": "Pemasaran Pariwisata",
    "500.13.3.1": "Pengembangan Pasar dan Informasi Pariwisata",
    "500.13.3.2": "Promosi Pariwisata Luar Negeri",
    "500.13.3.3": "Promosi Pariwisata Dalam Negeri",
    "500.13.3.4": "Pencitraan Indonesia",

    "500.13.4": "Ekonomi Kreatif Berbasis Seni dan Budaya",
    "500.13.4.1": "Pengembangan Industri Perfilman",
    "500.13.4.2": "Pengembangan Seni Pertunjukan dan Industri Musik",
    "500.13.4.3": "Pengembangan Seni Rupa",

    "500.13.5": "Ekonomi Kreatif Berbasis Media, Desain, dan IPTEK",
    "500.13.5.1": "Pengembangan Ekonomi Kreatif Berbasis Media",
    "500.13.5.2": "Desain dan Arsitektur",
    "500.13.5.3": "Kerjasama dan Fasilitasi",

    "500.13.6": "Pengembangan Sumber Daya Pariwisata dan Ekonomi Kreatif",
    "500.13.6.1": "Penelitian dan Pengembangan Kebijakan Kepariwisataan",
    "500.13.6.2": "Penelitian dan Pengembangan Kebijakan Ekonomi Kreatif",
    "500.13.6.3": "Pengembangan SDM Kepariwisataan dan Ekonomi Kreatif",
    "500.13.6.4": "Kompetensi Kepariwisataan dan Ekonomi Kreatif",
}

CATEGORY_CODES = [
    "500.13",
    "500.13.1",
    "500.13.2",
    "500.13.3",
    "500.13.4",
    "500.13.5",
    "500.13.6",
]

# Conversation states
WAIT_LETTER_DATE, WAIT_SUBJECT, WAIT_DESTINATION, CONFIRM = range(4)

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("bot-surat-keluar")


# ============================================================
# WAKTU & FORMAT TANGGAL
# ============================================================

MONTH_NAMES_ID = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}


def local_now() -> datetime:
    return datetime.now(LOCAL_TZ)


def now_iso() -> str:
    return local_now().isoformat(timespec="seconds")


def today_iso() -> str:
    return local_now().date().isoformat()


def format_date_id(value: str) -> str:
    try:
        return datetime.strptime(value, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        return value


def parse_letter_date(value: str) -> Optional[str]:
    value = value.strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def month_label(month_key: str) -> str:
    year, month = map(int, month_key.split("-"))
    return f"{MONTH_NAMES_ID[month]} {year}"


def shift_month(month_key: str, delta: int) -> str:
    year, month = map(int, month_key.split("-"))
    absolute = year * 12 + (month - 1) + delta
    new_year, month_index = divmod(absolute, 12)
    return f"{new_year:04d}-{month_index + 1:02d}"


def current_month_key() -> str:
    return local_now().strftime("%Y-%m")


# ============================================================
# DATABASE
# ============================================================

def ensure_db_parent() -> None:
    parent = Path(DB_PATH).expanduser().resolve().parent
    parent.mkdir(parents=True, exist_ok=True)


def db_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA foreign_keys=ON;")
    return conn


def column_exists(conn: sqlite3.Connection, table_name: str, column_name: str) -> bool:
    rows = conn.execute(f"PRAGMA table_info({table_name})").fetchall()
    return any(row["name"] == column_name for row in rows)


def init_db() -> None:
    ensure_db_parent()
    with db_connect() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS counters (
                scope TEXT PRIMARY KEY,
                current_number INTEGER NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS letters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sequence_number INTEGER NOT NULL,
                counter_scope TEXT NOT NULL,
                classification_code TEXT NOT NULL,
                classification_name TEXT NOT NULL,
                letter_number TEXT NOT NULL UNIQUE,
                letter_date TEXT NOT NULL,
                subject TEXT NOT NULL,
                destination TEXT,
                created_by INTEGER NOT NULL,
                created_by_name TEXT NOT NULL,
                created_at TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'ACTIVE'
            )
        """)

        # Migrasi otomatis database versi lama yang belum punya kolom tanggal surat.
        if not column_exists(conn, "letters", "letter_date"):
            conn.execute("ALTER TABLE letters ADD COLUMN letter_date TEXT")
            conn.execute("""
                UPDATE letters
                SET letter_date = substr(created_at, 1, 10)
                WHERE letter_date IS NULL OR letter_date = ''
            """)

        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_letters_created_at
            ON letters(created_at DESC)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_letters_classification
            ON letters(classification_code)
        """)
        conn.execute("""
            CREATE INDEX IF NOT EXISTS idx_letters_letter_date
            ON letters(letter_date)
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                action TEXT NOT NULL,
                detail TEXT NOT NULL,
                user_id INTEGER NOT NULL,
                user_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        conn.commit()


def get_counter_scope(classification_code: str) -> str:
    if SEQUENCE_MODE == "per_classification":
        return classification_code
    return "GLOBAL"


def allocate_letter_number(
    classification_code: str,
    letter_date: str,
    subject: str,
    destination: str,
    user_id: int,
    user_name: str,
) -> tuple[int, str]:
    """
    Alokasi nomor dilakukan dalam transaksi BEGIN IMMEDIATE agar dua pengguna
    tidak memperoleh nomor urut yang sama pada saat bersamaan.
    """
    scope = get_counter_scope(classification_code)
    conn = db_connect()
    try:
        conn.execute("BEGIN IMMEDIATE")

        row = conn.execute(
            "SELECT current_number FROM counters WHERE scope = ?",
            (scope,),
        ).fetchone()

        if row is None:
            next_number = START_NUMBER
            conn.execute(
                "INSERT INTO counters(scope, current_number) VALUES (?, ?)",
                (scope, next_number),
            )
        else:
            next_number = int(row["current_number"]) + 1
            conn.execute(
                "UPDATE counters SET current_number = ? WHERE scope = ?",
                (next_number, scope),
            )

        if next_number > 99999:
            raise ValueError("Nomor urut sudah melebihi batas 5 digit (99999).")

        letter_number = f"{classification_code}/{next_number:05d}"
        classification_name = CLASSIFICATIONS[classification_code]

        conn.execute("""
            INSERT INTO letters (
                sequence_number,
                counter_scope,
                classification_code,
                classification_name,
                letter_number,
                letter_date,
                subject,
                destination,
                created_by,
                created_by_name,
                created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            next_number,
            scope,
            classification_code,
            classification_name,
            letter_number,
            letter_date,
            subject,
            destination,
            user_id,
            user_name,
            now_iso(),
        ))

        conn.execute("""
            INSERT INTO audit_log(action, detail, user_id, user_name, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (
            "CREATE_NUMBER",
            f"{letter_number} | tanggal surat {letter_date}",
            user_id,
            user_name,
            now_iso(),
        ))

        conn.commit()
        return next_number, letter_number
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def get_last_letters(limit: int = 10) -> list[sqlite3.Row]:
    with db_connect() as conn:
        return conn.execute("""
            SELECT * FROM letters
            ORDER BY id DESC
            LIMIT ?
        """, (limit,)).fetchall()


def get_last_number() -> Optional[sqlite3.Row]:
    with db_connect() as conn:
        return conn.execute("""
            SELECT * FROM letters
            ORDER BY id DESC
            LIMIT 1
        """).fetchone()


def search_letters(term: str, limit: int = 15) -> list[sqlite3.Row]:
    pattern = f"%{term}%"
    with db_connect() as conn:
        return conn.execute("""
            SELECT * FROM letters
            WHERE letter_number LIKE ?
               OR letter_date LIKE ?
               OR subject LIKE ?
               OR destination LIKE ?
               OR classification_code LIKE ?
               OR classification_name LIKE ?
            ORDER BY id DESC
            LIMIT ?
        """, (
            pattern,
            pattern,
            pattern,
            pattern,
            pattern,
            pattern,
            limit,
        )).fetchall()


def get_letters_by_month(month_key: str) -> list[sqlite3.Row]:
    with db_connect() as conn:
        return conn.execute("""
            SELECT * FROM letters
            WHERE substr(letter_date, 1, 7) = ?
            ORDER BY letter_date ASC, id ASC
        """, (month_key,)).fetchall()


def get_max_sequence_global() -> int:
    with db_connect() as conn:
        row = conn.execute("""
            SELECT COALESCE(MAX(sequence_number), 0) AS max_seq
            FROM letters
            WHERE counter_scope = 'GLOBAL'
        """).fetchone()
        return int(row["max_seq"])


def set_next_global_number(next_number: int, user_id: int, user_name: str) -> None:
    if SEQUENCE_MODE != "global":
        raise ValueError("Perintah ini hanya tersedia saat SEQUENCE_MODE=global.")
    if not 1 <= next_number <= 99999:
        raise ValueError("Nomor berikutnya harus antara 1 dan 99999.")

    max_used = get_max_sequence_global()
    if next_number <= max_used:
        raise ValueError(
            f"Nomor berikutnya harus lebih besar dari nomor yang sudah pernah dipakai ({max_used:05d})."
        )

    current_number = next_number - 1
    with db_connect() as conn:
        conn.execute("""
            INSERT INTO counters(scope, current_number)
            VALUES ('GLOBAL', ?)
            ON CONFLICT(scope) DO UPDATE SET current_number = excluded.current_number
        """, (current_number,))
        conn.execute("""
            INSERT INTO audit_log(action, detail, user_id, user_name, created_at)
            VALUES (?, ?, ?, ?, ?)
        """, (
            "SET_NEXT_NUMBER",
            f"Next global number = {next_number:05d}",
            user_id,
            user_name,
            now_iso(),
        ))
        conn.commit()


def export_csv_file() -> str:
    fd, path = tempfile.mkstemp(prefix="surat_keluar_", suffix=".csv")
    os.close(fd)

    with db_connect() as conn:
        rows = conn.execute("""
            SELECT
                letter_date,
                letter_number,
                classification_code,
                classification_name,
                subject,
                destination,
                created_by_name,
                created_by,
                created_at,
                status
            FROM letters
            ORDER BY id ASC
        """).fetchall()

    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "Tanggal Surat",
            "Nomor Surat",
            "Kode Klasifikasi",
            "Nama Klasifikasi",
            "Perihal",
            "Tujuan",
            "Dibuat Oleh",
            "Telegram ID",
            "Waktu Input",
            "Status",
        ])
        for row in rows:
            writer.writerow([
                format_date_id(row["letter_date"]),
                row["letter_number"],
                row["classification_code"],
                row["classification_name"],
                row["subject"],
                row["destination"] or "",
                row["created_by_name"],
                row["created_by"],
                row["created_at"],
                row["status"],
            ])
    return path


# ============================================================
# LAPORAN BULANAN SPREADSHEET
# ============================================================

def build_monthly_xlsx(month_key: str) -> tuple[str, int]:
    rows = get_letters_by_month(month_key)

    wb = Workbook()
    ws = wb.active
    ws.title = "Register Surat"

    # Palet sederhana dan profesional.
    yellow = "F4D03F"
    dark = "1F2937"
    light_fill = "F8FAFC"
    border_color = "D1D5DB"
    white = "FFFFFF"

    thin = Side(style="thin", color=border_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Judul.
    ws.merge_cells("A1:J1")
    ws["A1"] = "LAPORAN BULANAN SURAT KELUAR BIDANG"
    ws["A1"].font = Font(size=16, bold=True, color=dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor=yellow)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Periode: {month_label(month_key)}"
    ws["A2"].font = Font(size=11, bold=True, color=dark)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:J3")
    ws["A3"] = f"Jumlah surat: {len(rows)} | Dibuat: {local_now().strftime('%d-%m-%Y %H:%M')} {APP_TIMEZONE}"
    ws["A3"].font = Font(size=10, italic=True, color="4B5563")
    ws["A3"].alignment = Alignment(horizontal="center")

    headers = [
        "No.",
        "Tanggal Surat",
        "Nomor Surat",
        "Kode Klasifikasi",
        "Klasifikasi",
        "Perihal",
        "Tujuan",
        "Dibuat Oleh",
        "Waktu Input",
        "Status",
    ]

    header_row = 5
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for index, row in enumerate(rows, start=1):
        excel_row = header_row + index
        try:
            letter_date_value = datetime.strptime(row["letter_date"], "%Y-%m-%d").date()
        except (TypeError, ValueError):
            letter_date_value = row["letter_date"]

        try:
            created_at_value = datetime.fromisoformat(row["created_at"]).replace(tzinfo=None)
        except (TypeError, ValueError):
            created_at_value = row["created_at"]

        values = [
            index,
            letter_date_value,
            row["letter_number"],
            row["classification_code"],
            row["classification_name"],
            row["subject"],
            row["destination"] or "-",
            row["created_by_name"],
            created_at_value,
            row["status"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if col_idx in {1, 2, 3, 4, 10} else "left",
                wrap_text=True,
            )
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=light_fill)

        ws.cell(excel_row, 2).number_format = "dd-mm-yyyy"
        ws.cell(excel_row, 9).number_format = "dd-mm-yyyy hh:mm"

    last_row = max(header_row + len(rows), header_row)

    if rows:
        table = Table(displayName="RegisterSuratBulanan", ref=f"A{header_row}:J{last_row}")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{header_row}:J{last_row}"
    ws.sheet_view.showGridLines = False

    widths = {
        "A": 6,
        "B": 15,
        "C": 24,
        "D": 18,
        "E": 42,
        "F": 42,
        "G": 30,
        "H": 24,
        "I": 20,
        "J": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Sheet rekap.
    recap = wb.create_sheet("Rekap Klasifikasi")
    recap.sheet_view.showGridLines = False
    recap.merge_cells("A1:D1")
    recap["A1"] = f"REKAP KLASIFIKASI — {month_label(month_key).upper()}"
    recap["A1"].font = Font(size=15, bold=True, color=dark)
    recap["A1"].fill = PatternFill("solid", fgColor=yellow)
    recap["A1"].alignment = Alignment(horizontal="center")
    recap.row_dimensions[1].height = 26

    recap["A3"] = "Total Surat"
    recap["B3"] = len(rows)
    recap["A3"].font = Font(bold=True)
    recap["B3"].font = Font(size=14, bold=True)

    recap_headers = ["No.", "Kode Klasifikasi", "Nama Klasifikasi", "Jumlah Surat"]
    for col_idx, value in enumerate(recap_headers, start=1):
        cell = recap.cell(row=5, column=col_idx, value=value)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    counts = Counter(row["classification_code"] for row in rows)
    sorted_counts = sorted(
        counts.items(),
        key=lambda item: (-item[1], item[0]),
    )

    for index, (code, count) in enumerate(sorted_counts, start=1):
        r = 5 + index
        values = [index, code, CLASSIFICATIONS.get(code, "-"), count]
        for col_idx, value in enumerate(values, start=1):
            cell = recap.cell(r, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                horizontal="center" if col_idx in {1, 2, 4} else "left",
                wrap_text=True,
            )
            if index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=light_fill)

    recap.column_dimensions["A"].width = 7
    recap.column_dimensions["B"].width = 20
    recap.column_dimensions["C"].width = 55
    recap.column_dimensions["D"].width = 16
    recap.freeze_panes = "A6"

    if sorted_counts:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Jumlah Surat per Klasifikasi"
        chart.y_axis.title = "Klasifikasi"
        chart.x_axis.title = "Jumlah Surat"
        chart.height = 8
        chart.width = 17

        data = Reference(
            recap,
            min_col=4,
            min_row=5,
            max_row=5 + len(sorted_counts),
        )
        categories = Reference(
            recap,
            min_col=2,
            min_row=6,
            max_row=5 + len(sorted_counts),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        recap.add_chart(chart, "F3")

    fd, path = tempfile.mkstemp(
        prefix=f"laporan_surat_keluar_{month_key}_",
        suffix=".xlsx",
    )
    os.close(fd)
    wb.save(path)
    return path, len(rows)


# ============================================================
# UTILITAS TELEGRAM
# ============================================================

def user_display_name(update: Update) -> str:
    user = update.effective_user
    if not user:
        return "Unknown"
    return user.full_name or user.username or str(user.id)


async def access_allowed(update: Update) -> bool:
    user = update.effective_user
    if not user:
        return False

    # Jika ALLOWED_USER_IDS kosong, bot terbuka untuk pengujian.
    if ALLOWED_USER_IDS and user.id not in ALLOWED_USER_IDS and user.id not in ADMIN_IDS:
        if update.callback_query:
            await update.callback_query.answer("Akses ditolak.", show_alert=True)
        elif update.effective_message:
            await update.effective_message.reply_text(
                "⛔ Anda belum terdaftar sebagai pengguna bot.\n\n"
                f"Telegram ID Anda: <code>{user.id}</code>\n"
                "Hubungi admin agar ID ini ditambahkan.",
                parse_mode=ParseMode.HTML,
            )
        return False
    return True


def is_admin(update: Update) -> bool:
    user = update.effective_user
    return bool(user and user.id in ADMIN_IDS)


def esc(text: Optional[str]) -> str:
    return html.escape(text or "-")


def short_label(text: str, max_len: int = 46) -> str:
    if len(text) <= max_len:
        return text
    return text[: max_len - 1] + "…"


def main_menu() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ Buat Nomor Surat", callback_data="menu:new")],
        [
            InlineKeyboardButton("🕘 Riwayat Terakhir", callback_data="menu:history"),
            InlineKeyboardButton("🔢 Nomor Terakhir", callback_data="menu:last"),
        ],
        [InlineKeyboardButton("📊 Laporan Bulanan", callback_data="menu:report")],
        [InlineKeyboardButton("🗂 Daftar Klasifikasi", callback_data="menu:classes")],
        [InlineKeyboardButton("🆔 ID Telegram Saya", callback_data="menu:myid")],
    ])


def category_keyboard() -> InlineKeyboardMarkup:
    rows = []
    for code in CATEGORY_CODES:
        rows.append([
            InlineKeyboardButton(
                f"{code} • {short_label(CLASSIFICATIONS[code], 34)}",
                callback_data=f"cat:{code}",
            )
        ])
    rows.append([InlineKeyboardButton("⬅️ Kembali", callback_data="menu:home")])
    return InlineKeyboardMarkup(rows)


def classification_keyboard(category_code: str) -> InlineKeyboardMarkup:
    rows = [[
        InlineKeyboardButton(
            f"✅ Gunakan {category_code}",
            callback_data=f"class:{category_code}",
        )
    ]]

    prefix = category_code + "."
    parent_depth = category_code.count(".")
    direct_children = []

    for code, name in CLASSIFICATIONS.items():
        if code.startswith(prefix) and code.count(".") == parent_depth + 1:
            direct_children.append((code, name))

    for code, name in direct_children:
        rows.append([
            InlineKeyboardButton(
                f"{code} • {short_label(name, 32)}",
                callback_data=f"class:{code}",
            )
        ])

    rows.append([InlineKeyboardButton("⬅️ Pilih Kelompok Lain", callback_data="menu:new")])
    return InlineKeyboardMarkup(rows)


def date_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(
            f"📅 Hari Ini ({local_now().strftime('%d-%m-%Y')})",
            callback_data="date:today",
        )],
        [InlineKeyboardButton("✍️ Masukkan Tanggal Lain", callback_data="date:manual")],
        [InlineKeyboardButton("❌ Batalkan", callback_data="confirm:no")],
    ])


def confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Terbitkan Nomor", callback_data="confirm:yes"),
            InlineKeyboardButton("❌ Batalkan", callback_data="confirm:no"),
        ]
    ])


def report_month_keyboard() -> InlineKeyboardMarkup:
    current = current_month_key()
    months = [shift_month(current, -i) for i in range(6)]

    rows = []
    for i in range(0, len(months), 2):
        row = []
        for key in months[i:i + 2]:
            row.append(
                InlineKeyboardButton(
                    month_label(key),
                    callback_data=f"report:{key}",
                )
            )
        rows.append(row)

    rows.append([InlineKeyboardButton("⬅️ Kembali", callback_data="menu:home")])
    return InlineKeyboardMarkup(rows)


# ============================================================
# HANDLERS
# ============================================================

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return

    text = (
        "📨 <b>BOT PENOMORAN SURAT KELUAR BIDANG</b>\n\n"
        "Format nomor:\n"
        "<code>KODE_KLASIFIKASI/00001</code>\n\n"
        "Contoh:\n"
        "<code>500.13.3.1/00001</code>\n\n"
        "Setiap register menyimpan tanggal surat, perihal, tujuan, pengguna, dan waktu input."
    )
    await update.effective_message.reply_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu(),
    )


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return

    text = (
        "ℹ️ <b>Panduan Singkat</b>\n\n"
        "• /start — buka menu utama\n"
        "• /baru — buat nomor surat baru\n"
        "• /riwayat — lihat 10 nomor terakhir\n"
        "• /terakhir — lihat nomor terakhir\n"
        "• /cari kata — cari nomor/perihal/tujuan/tanggal\n"
        "• /laporan — laporan spreadsheet bulan berjalan\n"
        "• /laporan 2026-07 — laporan spreadsheet bulan tertentu\n"
        "• /id — lihat Telegram ID\n"
        "• /batal — batalkan proses input\n\n"
        "<b>Admin:</b>\n"
        "• /setnomor 123 — set nomor berikutnya menjadi 00123\n"
        "• /export — ekspor seluruh register ke CSV"
    )
    await update.effective_message.reply_text(text, parse_mode=ParseMode.HTML)


async def my_id(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    if not user:
        return
    await update.effective_message.reply_text(
        f"🆔 Telegram ID Anda: <code>{user.id}</code>",
        parse_mode=ParseMode.HTML,
    )


async def menu_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> Optional[int]:
    if not await access_allowed(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "menu:home":
        await query.edit_message_text(
            "📨 <b>BOT PENOMORAN SURAT KELUAR BIDANG</b>\n\nSilakan pilih menu:",
            parse_mode=ParseMode.HTML,
            reply_markup=main_menu(),
        )
        return ConversationHandler.END

    if data == "menu:new":
        context.user_data.clear()
        await query.edit_message_text(
            "🗂 <b>Pilih kelompok klasifikasi surat:</b>",
            parse_mode=ParseMode.HTML,
            reply_markup=category_keyboard(),
        )
        return ConversationHandler.END

    if data == "menu:history":
        await show_history_message(query.edit_message_text)
        return ConversationHandler.END

    if data == "menu:last":
        await show_last_message(query.edit_message_text)
        return ConversationHandler.END

    if data == "menu:report":
        await query.edit_message_text(
            "📊 <b>Pilih bulan laporan spreadsheet:</b>\n\n"
            "Laporan dibuat berdasarkan <b>tanggal surat keluar</b>.",
            parse_mode=ParseMode.HTML,
            reply_markup=report_month_keyboard(),
        )
        return ConversationHandler.END

    if data == "menu:classes":
        await query.edit_message_text(
            build_classification_list(),
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Kembali", callback_data="menu:home")]
            ]),
        )
        return ConversationHandler.END

    if data == "menu:myid":
        user = update.effective_user
        await query.edit_message_text(
            f"🆔 Telegram ID Anda: <code>{user.id}</code>",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Kembali", callback_data="menu:home")]
            ]),
        )
        return ConversationHandler.END

    return ConversationHandler.END


async def category_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return

    query = update.callback_query
    await query.answer()
    code = query.data.split(":", 1)[1]

    await query.edit_message_text(
        f"🗂 <b>{esc(code)} — {esc(CLASSIFICATIONS[code])}</b>\n\n"
        "Pilih kode yang akan digunakan:",
        parse_mode=ParseMode.HTML,
        reply_markup=classification_keyboard(code),
    )


async def new_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    context.user_data.clear()
    await update.effective_message.reply_text(
        "🗂 <b>Pilih kelompok klasifikasi surat:</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=category_keyboard(),
    )
    return ConversationHandler.END


async def class_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()

    code = query.data.split(":", 1)[1]
    context.user_data["classification_code"] = code
    context.user_data["classification_name"] = CLASSIFICATIONS[code]

    await query.edit_message_text(
        f"✅ <b>Klasifikasi dipilih</b>\n"
        f"<code>{esc(code)}</code>\n"
        f"{esc(CLASSIFICATIONS[code])}\n\n"
        "📅 Pilih <b>tanggal surat keluar</b>:",
        parse_mode=ParseMode.HTML,
        reply_markup=date_keyboard(),
    )
    return WAIT_LETTER_DATE


async def use_today_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()
    context.user_data["letter_date"] = today_iso()

    await query.edit_message_text(
        f"📅 <b>Tanggal surat:</b> {format_date_id(context.user_data['letter_date'])}\n\n"
        "📝 Ketik <b>perihal surat</b>:",
        parse_mode=ParseMode.HTML,
    )
    return WAIT_SUBJECT


async def manual_date_prompt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()
    await query.edit_message_text(
        "✍️ Ketik <b>tanggal surat keluar</b> dengan format:\n"
        "<code>DD-MM-YYYY</code>\n\n"
        "Contoh:\n"
        "<code>13-07-2026</code>",
        parse_mode=ParseMode.HTML,
    )
    return WAIT_LETTER_DATE


async def receive_letter_date(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    raw = (update.effective_message.text or "").strip()
    parsed = parse_letter_date(raw)

    if not parsed:
        await update.effective_message.reply_text(
            "⚠️ Format tanggal tidak valid.\n\n"
            "Gunakan format <code>DD-MM-YYYY</code>, contoh <code>13-07-2026</code>.",
            parse_mode=ParseMode.HTML,
        )
        return WAIT_LETTER_DATE

    context.user_data["letter_date"] = parsed
    await update.effective_message.reply_text(
        f"📅 <b>Tanggal surat:</b> {format_date_id(parsed)}\n\n"
        "📝 Ketik <b>perihal surat</b>:",
        parse_mode=ParseMode.HTML,
    )
    return WAIT_SUBJECT


async def receive_subject(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    subject = (update.effective_message.text or "").strip()
    if len(subject) < 3:
        await update.effective_message.reply_text(
            "Perihal terlalu pendek. Silakan ketik perihal surat yang lebih jelas."
        )
        return WAIT_SUBJECT

    context.user_data["subject"] = subject

    await update.effective_message.reply_text(
        "🎯 Ketik <b>tujuan/penerima surat</b>.\n"
        "Atau tekan <b>Lewati</b> jika tidak ingin dicatat.",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⏭ Lewati", callback_data="destination:skip")]
        ]),
    )
    return WAIT_DESTINATION


async def skip_destination(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()
    context.user_data["destination"] = "-"
    await show_confirmation(query.edit_message_text, context)
    return CONFIRM


async def receive_destination(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    destination = (update.effective_message.text or "").strip()
    context.user_data["destination"] = destination or "-"
    await show_confirmation(update.effective_message.reply_text, context)
    return CONFIRM


async def show_confirmation(send_func, context: ContextTypes.DEFAULT_TYPE) -> None:
    code = context.user_data["classification_code"]
    name = context.user_data["classification_name"]
    letter_date = context.user_data["letter_date"]
    subject = context.user_data["subject"]
    destination = context.user_data.get("destination", "-")

    text = (
        "🔎 <b>Konfirmasi Data Surat</b>\n\n"
        f"<b>Tanggal surat:</b> {esc(format_date_id(letter_date))}\n\n"
        f"<b>Klasifikasi:</b>\n<code>{esc(code)}</code>\n{esc(name)}\n\n"
        f"<b>Perihal:</b>\n{esc(subject)}\n\n"
        f"<b>Tujuan:</b>\n{esc(destination)}\n\n"
        "Nomor 5 digit akan diterbitkan saat tombol konfirmasi ditekan."
    )
    await send_func(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=confirm_keyboard(),
    )


async def confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    if not await access_allowed(update):
        return ConversationHandler.END

    query = update.callback_query
    await query.answer()
    choice = query.data

    if choice == "confirm:no":
        context.user_data.clear()
        await query.edit_message_text(
            "❌ Pembuatan nomor dibatalkan.",
            reply_markup=main_menu(),
        )
        return ConversationHandler.END

    required = {
        "classification_code",
        "classification_name",
        "letter_date",
        "subject",
    }
    if not required.issubset(context.user_data):
        context.user_data.clear()
        await query.edit_message_text(
            "⚠️ Data proses sudah tidak lengkap. Silakan mulai lagi.",
            reply_markup=main_menu(),
        )
        return ConversationHandler.END

    code = context.user_data["classification_code"]
    letter_date = context.user_data["letter_date"]
    subject = context.user_data["subject"]
    destination = context.user_data.get("destination", "-")
    user = update.effective_user

    try:
        sequence, letter_number = await asyncio.to_thread(
            allocate_letter_number,
            code,
            letter_date,
            subject,
            destination,
            user.id,
            user_display_name(update),
        )
    except sqlite3.IntegrityError:
        logger.exception("Duplicate letter number detected")
        await query.edit_message_text(
            "⚠️ Terjadi benturan nomor. Silakan coba lagi.",
            reply_markup=main_menu(),
        )
        return ConversationHandler.END
    except Exception as exc:
        logger.exception("Failed to allocate number")
        await query.edit_message_text(
            f"⚠️ Gagal membuat nomor: {esc(str(exc))}",
            parse_mode=ParseMode.HTML,
            reply_markup=main_menu(),
        )
        return ConversationHandler.END

    text = (
        "✅ <b>NOMOR SURAT BERHASIL DITERBITKAN</b>\n\n"
        f"📌 <b>Nomor:</b>\n"
        f"<code>{esc(letter_number)}</code>\n\n"
        f"📅 <b>Tanggal Surat:</b>\n"
        f"{esc(format_date_id(letter_date))}\n\n"
        f"🗂 <b>Klasifikasi:</b>\n"
        f"{esc(CLASSIFICATIONS[code])}\n\n"
        f"📝 <b>Perihal:</b>\n"
        f"{esc(subject)}\n\n"
        f"🎯 <b>Tujuan:</b>\n"
        f"{esc(destination)}\n\n"
        f"👤 <b>Dibuat oleh:</b> {esc(user_display_name(update))}\n"
        f"🕒 <b>Waktu input:</b> {esc(local_now().strftime('%d-%m-%Y %H:%M:%S'))}"
    )

    context.user_data.clear()
    await query.edit_message_text(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=main_menu(),
    )
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.effective_message.reply_text(
        "❌ Proses dibatalkan.",
        reply_markup=main_menu(),
    )
    return ConversationHandler.END


def format_letter_row(row: sqlite3.Row) -> str:
    created = row["created_at"]
    try:
        created_dt = datetime.fromisoformat(created)
        created_text = created_dt.strftime("%d-%m-%Y %H:%M")
    except (TypeError, ValueError):
        created_text = created

    return (
        f"📌 <code>{esc(row['letter_number'])}</code>\n"
        f"📅 {esc(format_date_id(row['letter_date']))}\n"
        f"📝 {esc(row['subject'])}\n"
        f"🎯 {esc(row['destination'] or '-')}\n"
        f"👤 {esc(row['created_by_name'])} • {esc(created_text)}"
    )


async def show_history_message(send_func) -> None:
    rows = await asyncio.to_thread(get_last_letters, 10)
    if not rows:
        text = "📭 Belum ada nomor surat yang tersimpan."
    else:
        parts = ["🕘 <b>10 NOMOR SURAT TERAKHIR</b>"]
        for i, row in enumerate(rows, start=1):
            parts.append(f"\n<b>{i}.</b> {format_letter_row(row)}")
        text = "\n".join(parts)

    await send_func(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Kembali", callback_data="menu:home")]
        ]),
    )


async def history_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return
    await show_history_message(update.effective_message.reply_text)


async def show_last_message(send_func) -> None:
    row = await asyncio.to_thread(get_last_number)
    if not row:
        text = "📭 Belum ada nomor surat yang tersimpan."
    else:
        text = (
            "🔢 <b>NOMOR SURAT TERAKHIR</b>\n\n"
            + format_letter_row(row)
        )

    await send_func(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⬅️ Kembali", callback_data="menu:home")]
        ]),
    )


async def last_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return
    await show_last_message(update.effective_message.reply_text)


async def search_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return

    term = " ".join(context.args).strip()
    if not term:
        await update.effective_message.reply_text(
            "Gunakan format:\n<code>/cari kata_kunci</code>\n\n"
            "Contoh:\n<code>/cari promosi</code>\n"
            "<code>/cari 500.13.3.1</code>\n"
            "<code>/cari 00025</code>\n"
            "<code>/cari 2026-07-13</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    rows = await asyncio.to_thread(search_letters, term, 15)
    if not rows:
        await update.effective_message.reply_text(
            f"🔎 Tidak ada hasil untuk: <b>{esc(term)}</b>",
            parse_mode=ParseMode.HTML,
        )
        return

    parts = [f"🔎 <b>HASIL PENCARIAN: {esc(term)}</b>"]
    for i, row in enumerate(rows, start=1):
        parts.append(f"\n<b>{i}.</b> {format_letter_row(row)}")

    await update.effective_message.reply_text(
        "\n".join(parts),
        parse_mode=ParseMode.HTML,
    )


def build_classification_list() -> str:
    lines = ["🗂 <b>DAFTAR KLASIFIKASI</b>\n"]
    for code, name in CLASSIFICATIONS.items():
        lines.append(f"<code>{esc(code)}</code> — {esc(name)}")
    return "\n".join(lines)


async def classes_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return
    await update.effective_message.reply_text(
        build_classification_list(),
        parse_mode=ParseMode.HTML,
    )


async def send_monthly_report(message, month_key: str) -> None:
    path, count = await asyncio.to_thread(build_monthly_xlsx, month_key)
    try:
        with open(path, "rb") as f:
            await message.reply_document(
                document=f,
                filename=f"laporan_surat_keluar_{month_key}.xlsx",
                caption=(
                    f"📊 <b>Laporan Surat Keluar — {esc(month_label(month_key))}</b>\n"
                    f"Jumlah surat: <b>{count}</b>\n\n"
                    "Isi file:\n"
                    "• Register Surat\n"
                    "• Rekap Klasifikasi"
                ),
                parse_mode=ParseMode.HTML,
            )
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


async def report_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return

    month_key = current_month_key()
    if context.args:
        month_key = context.args[0].strip()

    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month_key):
        await update.effective_message.reply_text(
            "⚠️ Format bulan tidak valid.\n\n"
            "Gunakan:\n"
            "<code>/laporan 2026-07</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    await update.effective_message.reply_text(
        f"⏳ Menyiapkan laporan <b>{esc(month_label(month_key))}</b>...",
        parse_mode=ParseMode.HTML,
    )
    await send_monthly_report(update.effective_message, month_key)


async def report_callback(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return

    query = update.callback_query
    await query.answer("Menyiapkan laporan...")
    month_key = query.data.split(":", 1)[1]

    if not re.fullmatch(r"\d{4}-(0[1-9]|1[0-2])", month_key):
        await query.edit_message_text("⚠️ Bulan laporan tidak valid.")
        return

    await query.edit_message_text(
        f"⏳ Menyiapkan spreadsheet <b>{esc(month_label(month_key))}</b>...",
        parse_mode=ParseMode.HTML,
    )
    await send_monthly_report(query.message, month_key)
    await query.message.reply_text(
        "✅ Laporan selesai dibuat.",
        reply_markup=main_menu(),
    )


async def set_number_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return
    if not is_admin(update):
        await update.effective_message.reply_text("⛔ Perintah ini hanya untuk admin.")
        return

    if len(context.args) != 1 or not context.args[0].isdigit():
        await update.effective_message.reply_text(
            "Format:\n<code>/setnomor 123</code>\n\n"
            "Artinya nomor berikutnya yang diterbitkan adalah <code>00123</code>.",
            parse_mode=ParseMode.HTML,
        )
        return

    next_number = int(context.args[0])
    user = update.effective_user

    try:
        await asyncio.to_thread(
            set_next_global_number,
            next_number,
            user.id,
            user_display_name(update),
        )
    except Exception as exc:
        await update.effective_message.reply_text(
            f"⚠️ Gagal: {esc(str(exc))}",
            parse_mode=ParseMode.HTML,
        )
        return

    await update.effective_message.reply_text(
        f"✅ Nomor berikutnya diset menjadi <code>{next_number:05d}</code>.",
        parse_mode=ParseMode.HTML,
    )


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await access_allowed(update):
        return
    if not is_admin(update):
        await update.effective_message.reply_text("⛔ Perintah ini hanya untuk admin.")
        return

    path = await asyncio.to_thread(export_csv_file)
    try:
        with open(path, "rb") as f:
            await update.effective_message.reply_document(
                document=f,
                filename=f"register_surat_keluar_{local_now().strftime('%Y%m%d_%H%M%S')}.csv",
                caption="📊 Export seluruh register nomor surat keluar.",
            )
    finally:
        try:
            os.remove(path)
        except OSError:
            pass


async def unknown_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.effective_message.reply_text(
        "Perintah tidak dikenali. Ketik /help untuk melihat panduan."
    )


async def post_init(application: Application) -> None:
    commands = [
        BotCommand("start", "Buka menu utama"),
        BotCommand("baru", "Buat nomor surat baru"),
        BotCommand("riwayat", "Lihat 10 nomor terakhir"),
        BotCommand("terakhir", "Lihat nomor terakhir"),
        BotCommand("cari", "Cari register surat"),
        BotCommand("laporan", "Download laporan bulanan Excel"),
        BotCommand("klasifikasi", "Lihat daftar klasifikasi"),
        BotCommand("id", "Lihat Telegram ID"),
        BotCommand("help", "Panduan penggunaan"),
        BotCommand("batal", "Batalkan proses input"),
    ]
    await application.bot.set_my_commands(commands)


def build_application() -> Application:
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN belum diisi.")

    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    conversation = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(class_callback, pattern=r"^class:"),
        ],
        states={
            WAIT_LETTER_DATE: [
                CallbackQueryHandler(use_today_date, pattern=r"^date:today$"),
                CallbackQueryHandler(manual_date_prompt, pattern=r"^date:manual$"),
                CallbackQueryHandler(confirm_callback, pattern=r"^confirm:no$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_letter_date),
            ],
            WAIT_SUBJECT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_subject),
            ],
            WAIT_DESTINATION: [
                CallbackQueryHandler(skip_destination, pattern=r"^destination:skip$"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_destination),
            ],
            CONFIRM: [
                CallbackQueryHandler(confirm_callback, pattern=r"^confirm:"),
            ],
        },
        fallbacks=[
            CommandHandler("batal", cancel),
            CommandHandler("start", start),
        ],
        allow_reentry=True,
    )

    # Command handlers
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("id", my_id))
    app.add_handler(CommandHandler("baru", new_command))
    app.add_handler(CommandHandler("riwayat", history_command))
    app.add_handler(CommandHandler("terakhir", last_command))
    app.add_handler(CommandHandler("cari", search_command))
    app.add_handler(CommandHandler("laporan", report_command))
    app.add_handler(CommandHandler("klasifikasi", classes_command))
    app.add_handler(CommandHandler("setnomor", set_number_command))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("batal", cancel))

    # Conversation harus dipasang sebelum generic callback handlers.
    app.add_handler(conversation)

    # Callback handlers
    app.add_handler(CallbackQueryHandler(menu_callback, pattern=r"^menu:"))
    app.add_handler(CallbackQueryHandler(category_callback, pattern=r"^cat:"))
    app.add_handler(CallbackQueryHandler(report_callback, pattern=r"^report:"))

    # Unknown commands
    app.add_handler(MessageHandler(filters.COMMAND, unknown_command))

    return app


def main() -> None:
    init_db()
    app = build_application()
    logger.info(
        "Bot started. DB=%s | SEQUENCE_MODE=%s | TIMEZONE=%s",
        DB_PATH,
        SEQUENCE_MODE,
        APP_TIMEZONE,
    )
    app.run_polling(
        allowed_updates=Update.ALL_TYPES,
        drop_pending_updates=False,
    )


if __name__ == "__main__":
    main()
